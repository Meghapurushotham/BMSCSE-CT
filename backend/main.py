from backend.database import init_db, log_action, get_db
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import List
import os, uuid
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from paddleocr import PaddleOCR, PPStructure
from pdf2image import convert_from_path
from pydantic import BaseModel

from PIL import Image
import numpy as np

# ==================== APP ====================
app = FastAPI(title="BMSCSE-CT Phase-2.4 Stable")

@app.on_event("startup")
def startup():
    init_db()

# ==================== CORS ====================
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== DIRECTORIES ====================
UPLOAD_DIR = "uploads"
OUTPUT_DIR = "outputs"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ==================== OCR ENGINES ====================
ocr = PaddleOCR(use_angle_cls=False, lang="en", show_log=False)
table_engine = PPStructure(show_log=False, lang="en")

# ==================== MODELS ====================
class LoginRequest(BaseModel):
    email: str

class AgentTask(BaseModel):
    task: str

# ==================== IMAGE SAFETY ====================
def load_image_safe(img_or_path):
    img = Image.open(img_or_path) if isinstance(img_or_path, str) else img_or_path
    if img.mode != "RGB":
        img = img.convert("RGB")
    return np.array(img)

# ==================== SAFE TEXT EXTRACTION ====================
def extract_text_safe(res):
    texts = []
    if isinstance(res, dict) and "text" in res:
        texts.append(res["text"])
    elif isinstance(res, list):
        for item in res:
            if isinstance(item, dict) and "text" in item:
                texts.append(item["text"])
    return texts

# ==================== HOME ====================
@app.get("/")
def home():
    return {"message": "BMSCSE-CT backend running successfully"}

# ==================== EXCEL FORMAT ====================
def format_excel_table(excel_path):
    wb = load_workbook(excel_path)

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            ws.append(["No data detected"])

        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin

        for col in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = length + 3

        ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
        table = Table(displayName=f"Table_{ws.title}", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True
        )
        ws.add_table(table)
        ws.freeze_panes = "A2"

    wb.save(excel_path)

# ==================== LOGIN ====================
@app.post("/login")
def login(data: LoginRequest):
    email = data.email.lower()

    if not email.endswith("@bmsce.ac.in"):
        raise HTTPException(status_code=403, detail="Only BMSCE email allowed")

    if email.startswith("principal"):
        role = "principal"
    elif email.startswith("hod"):
        role = "hod"
    elif email.startswith("admin"):
        role = "admin"
    elif "student" in email:
        role = "student"
    else:
        role = "user"

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT OR IGNORE INTO users (email, role) VALUES (?, ?)",
        (email, role)
    )
    conn.commit()
    conn.close()

    return {"status": "success", "email": email, "role": role}

# ==================== AGENT ROUTER ====================
@app.post("/agent/route-task")
def route_task(data: AgentTask):
    task = data.task.lower()
    advanced_keywords = [
        "handwritten", "unclear", "multiple tables",
        "many pages", "scanned", "complex",
        "merged cells", "symbols"
    ]

    for word in advanced_keywords:
        if word in task:
            return {
                "mode": "advanced",
                "endpoint": "/convert-advanced/",
                "message": "Advanced Document Agent selected"
            }

    return {
        "mode": "standard",
        "endpoint": "/convert-multiple/",
        "message": "Standard Document Agent selected"
    }

# ==================== STANDARD CONVERT ====================
@app.post("/convert-multiple/")
def convert_multiple(files: List[UploadFile] = File(...)):
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")

    file_id = str(uuid.uuid4())
    output_excel = os.path.join(OUTPUT_DIR, f"{file_id}.xlsx")

    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        for idx, file in enumerate(files):
            input_path = os.path.join(UPLOAD_DIR, f"{file_id}_{file.filename}")
            with open(input_path, "wb") as f:
                f.write(file.file.read())

            sheet_name = f"File_{idx+1}"
            final_df = None

            try:
                if input_path.lower().endswith(".pdf"):
                    images = convert_from_path(input_path, first_page=1, last_page=2)
                else:
                    images = [Image.open(input_path)]

                tables = []
                for img in images:
                    result = table_engine(load_image_safe(img))
                    for block in result:
                        if block.get("type") == "table":
                            tables.extend(pd.read_html(block["res"]["html"]))

                if tables:
                    final_df = pd.concat(tables, ignore_index=True)

            except:
                pass

            if final_df is None or final_df.empty:
                ocr_res = ocr.ocr(input_path, cls=False)
                rows = [[l[1][0]] for l in ocr_res[0]] if ocr_res else [["No text"]]
                final_df = pd.DataFrame(rows, columns=["Extracted Text"])

            final_df.to_excel(writer, sheet_name=sheet_name, index=False)

    log_action(
        files[0].headers.get("x-user-email", "unknown@bmsce.ac.in"),
        f"Converted {len(files)} file(s)"
    )

    format_excel_table(output_excel)

    return {"status": "success", "file_id": file_id}

# ==================== ADVANCED CONVERT (FIXED) ====================
@app.post("/convert-advanced/")
def convert_advanced(files: List[UploadFile] = File(...)):
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")

    file_id = str(uuid.uuid4())
    output_excel = os.path.join(OUTPUT_DIR, f"ADV_{file_id}.xlsx")
    wrote_sheet = False

    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        for f_idx, file in enumerate(files):
            input_path = os.path.join(UPLOAD_DIR, f"{file_id}_{file.filename}")
            with open(input_path, "wb") as f:
                f.write(file.file.read())

            pages = convert_from_path(input_path) if input_path.endswith(".pdf") else [Image.open(input_path)]

            for p_idx, page in enumerate(pages):
                layout = table_engine(load_image_safe(page))
                text_rows = []
                table_count = 0

                for block in layout:
                    if block.get("type") == "table":
                        try:
                            for df in pd.read_html(block["res"]["html"]):
                                table_count += 1
                                df.to_excel(
                                    writer,
                                    sheet_name=f"F{f_idx+1}_P{p_idx+1}_T{table_count}",
                                    index=False
                                )
                                wrote_sheet = True
                        except:
                            pass
                    else:
                        for txt in extract_text_safe(block.get("res")):
                            text_rows.append([txt])

                if text_rows:
                    pd.DataFrame(text_rows, columns=["Detected Text"]).to_excel(
                        writer,
                        sheet_name=f"F{f_idx+1}_P{p_idx+1}_TEXT",
                        index=False
                    )
                    wrote_sheet = True

        if not wrote_sheet:
            pd.DataFrame([["No readable data detected"]]).to_excel(
                writer, sheet_name="EMPTY", index=False
            )

    format_excel_table(output_excel)

    return {
        "status": "success",
        "message": "Advanced multi-page, multi-table conversion completed",
        "file_id": f"ADV_{file_id}"
    }

# ==================== DOWNLOAD ====================
@app.get("/download/{file_id}")
def download_excel(file_id: str):
    path = os.path.join(OUTPUT_DIR, f"{file_id}.xlsx")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(
        path,
        filename="BMSCSE_CT_Converted.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
